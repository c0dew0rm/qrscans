import openpyxl
import pyqrcode

xlsFilePath = '/home/rahul/gitlab/QRGeneration/QR Code Label Rajastan.xlsx'
wb_obj = openpyxl.load_workbook(xlsFilePath)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# text_file = open("sample.txt", "w")
def split(word): 
    return [char for char in word]

def convert(s): 
    str1 = "" 
    return(str1.join(s)) 

for i in range(2, max_row + 1):
    qrString = ""
    for j in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = i, column = j)
        if(j < max_col):
            if(j == 5):
                datestring = str(cell_obj.value)
                datestringArr = datestring.split('.')
                year = datestringArr[2]
                yearArr = split(year)
                yearLastTwo = yearArr[2]+yearArr[3]
                datestring = datestringArr[0] + '.' + datestring[1] + '.' + yearLastTwo
                qrString = qrString + "DOT:" + datestring +"\n"
            elif(j == 6):
                datestring = str(cell_obj.value)
                datestringArr = datestring.split('.')
                year = datestringArr[2]
                yearArr = split(year)
                yearLastTwo = yearArr[2]+yearArr[3]
                datestring = datestringArr[0] + '.' + datestring[1] + '.' + yearLastTwo
                qrString = qrString + "DOP:" + datestring +"\n"
            elif(j == 7):
                datestring = str(cell_obj.value)
                datestringArr = datestring.split('.')
                year = datestringArr[2]
                yearArr = split(year)
                yearLastTwo = yearArr[2]+yearArr[3]
                datestring = datestringArr[0] + '.' + datestring[1] + '.' + yearLastTwo
                qrString = qrString + "EXP:" + datestring +"\n"
            else:
                qrString = qrString + str(cell_obj.value) +"\n"
        else:
            qrString = qrString + str(cell_obj.value)+"\n"
            qrCode = pyqrcode.create(qrString,error='L')
            imgName = "10"
            offset = "000000"
            offsetArr = split(offset)
            rowString = str(i-1)
            offsetIndex = len(offset) - 1
            for k in range(len(rowString)-1, -1, -1):
                offsetArr[offsetIndex] = rowString[k]
                offsetIndex -= 1
            imgName = imgName + convert(offsetArr) + '_'
            qrCode.png(imgName,scale = 6, quiet_zone = 0)
    qrString = ""